"""
Excel Reporter (Overhauled)

Generates a single Excel workbook with all elimination, evaluation,
validation, and configuration details for the model development pipeline.
"""

from typing import Any, Dict, List, Optional
from datetime import datetime
from pathlib import Path
import logging

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

from src.config.schema import PipelineConfig

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
KEPT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ELIM_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
PASS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
FAIL_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class ExcelReporter:
    """Generates the full model development Excel report.

    Keeps the original 11 sheets with styling improvements and adds new
    sheets for validation results, score distribution, and config dump.
    """

    def __init__(self, config: Optional[PipelineConfig] = None):
        self.config = config

    # ======================================================================
    # Public API
    # ======================================================================

    def generate(self, output_path: str, pipeline_results: Dict[str, Any]) -> str:
        """Generate the full Excel report from pipeline results.

        Args:
            output_path: File path for the .xlsx output.
            pipeline_results: Dict with keys matching the expected structure:
                summary, step_results, corr_pairs_df, selection_df,
                performance_df, lift_tables, importance_df,
                validation_report, score_distribution, config.

        Returns:
            The output_path written.
        """
        wb = Workbook()

        # 00_Summary
        self._write_summary_sheet(
            wb,
            pipeline_results.get("summary", {}),
            pipeline_results.get("validation_report"),
        )

        # 01-04: Elimination sheets from step_results
        for step_result in pipeline_results.get("step_results", []):
            df = step_result.get("results_df")
            name = step_result.get("step_name", "Step")
            if df is not None:
                self._write_df_sheet(wb, name, df)

        # 05_Corr_Pairs
        corr_pairs = pipeline_results.get("corr_pairs_df")
        if corr_pairs is not None and len(corr_pairs) > 0:
            self._write_df_sheet(wb, "05_Corr_Pairs", corr_pairs)

        # 06_Selection
        selection_df = pipeline_results.get("selection_df")
        if selection_df is not None:
            self._write_df_sheet(wb, "06_Selection", selection_df)

        # 07_Performance
        perf_df = pipeline_results.get("performance_df")
        if perf_df is not None:
            self._write_df_sheet(wb, "07_Performance", perf_df)

        # 07_Lift_Tables
        lift_tables = pipeline_results.get("lift_tables")
        if lift_tables:
            self._write_lift_sheets(wb, lift_tables)

        # 07_Importance
        importance_df = pipeline_results.get("importance_df")
        if importance_df is not None:
            self._write_df_sheet(wb, "07_Importance", importance_df)

        # 08_Validation (NEW)
        validation_report = pipeline_results.get("validation_report")
        if validation_report is not None:
            self._write_validation_sheet(wb, validation_report)

        # 09_Score_Distribution (NEW)
        score_dist = pipeline_results.get("score_distribution")
        if score_dist:
            self._write_score_distribution_sheet(wb, score_dist)

        # 10_Config (NEW)
        config_obj = pipeline_results.get("config") or self.config
        if config_obj is not None:
            self._write_config_sheet(wb, config_obj)

        # Remove the default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("COMPLETE | Excel saved: %s", output_path)
        return output_path

    # ======================================================================
    # Sheet writers
    # ======================================================================

    def _write_summary_sheet(
        self,
        wb: Workbook,
        summary: Dict[str, Any],
        validation_report: Any = None,
    ) -> None:
        """Write the 00_Summary sheet."""
        ws = wb.create_sheet("00_Summary")
        ws.sheet_properties.tabColor = "2F5496"

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 55

        # Header comment
        ws["A1"].comment = Comment(
            "Pipeline summary: key metrics, data stats, and validation flags.", "ExcelReporter"
        )

        # Title
        ws["A1"] = "Credit Scoring Model Development Report"
        ws["A1"].font = Font(bold=True, size=14, color="2F5496")
        ws.merge_cells("A1:B1")

        row = 3
        for key, value in summary.items():
            cell_a = ws.cell(row=row, column=1, value=key)
            cell_b = ws.cell(row=row, column=2, value=str(value))
            cell_a.font = Font(bold=True)
            cell_a.border = THIN_BORDER
            cell_b.border = THIN_BORDER
            row += 1

        # Validation summary block
        if validation_report is not None:
            row += 1
            ws.cell(row=row, column=1, value="Validation Results").font = Font(
                bold=True, size=12, color="2F5496"
            )
            row += 1

            # Try to get counts from the report
            pass_count = getattr(validation_report, "pass_count", 0)
            warn_count = getattr(validation_report, "warning_count", 0)
            fail_count = getattr(validation_report, "fail_count", 0)
            has_crit = getattr(validation_report, "has_critical_failures", False)

            for label, val, fill in [
                ("Checks Passed", pass_count, PASS_FILL),
                ("Checks Warning", warn_count, WARN_FILL),
                ("Checks Failed", fail_count, FAIL_FILL),
            ]:
                ca = ws.cell(row=row, column=1, value=label)
                cb = ws.cell(row=row, column=2, value=val)
                ca.font = Font(bold=True)
                ca.border = THIN_BORDER
                cb.border = THIN_BORDER
                cb.fill = fill
                row += 1

            status_cell = ws.cell(
                row=row, column=1, value="Critical Failures"
            )
            val_cell = ws.cell(
                row=row, column=2,
                value="YES -- Review 08_Validation sheet" if has_crit else "None",
            )
            status_cell.font = Font(bold=True)
            status_cell.border = THIN_BORDER
            val_cell.border = THIN_BORDER
            val_cell.fill = FAIL_FILL if has_crit else PASS_FILL

    def _write_df_sheet(
        self, wb: Workbook, sheet_name: str, df: pd.DataFrame
    ) -> None:
        """Write a DataFrame to a styled sheet."""
        if df is None or len(df) == 0:
            ws = wb.create_sheet(sheet_name)
            ws["A1"] = "No data"
            return

        ws = wb.create_sheet(sheet_name)

        # Sheet description comment
        ws["A1"].comment = Comment(
            f"Data for pipeline step: {sheet_name}", "ExcelReporter"
        )

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Write data
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER
                cell.value = _safe_value(value)

            # Color-code by Status column
            if "Status" in df.columns:
                status_col = list(df.columns).index("Status") + 1
                status_val = ws.cell(row=row_idx, column=status_col).value
                fill = None
                if status_val == "Eliminated":
                    fill = ELIM_FILL
                elif status_val in ("Kept", "Added"):
                    fill = KEPT_FILL
                if fill:
                    for ci in range(1, len(df.columns) + 1):
                        ws.cell(row=row_idx, column=ci).fill = fill

        _autofit_columns(ws, df)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_lift_sheets(
        self, wb: Workbook, lift_tables: Dict[str, pd.DataFrame]
    ) -> None:
        """Write combined lift table sheet."""
        if not lift_tables:
            return
        combined_rows = []
        for period, lt in lift_tables.items():
            lt_copy = lt.copy()
            lt_copy.insert(0, "Period", period)
            combined_rows.append(lt_copy)
        if combined_rows:
            combined = pd.concat(combined_rows, ignore_index=True)
            self._write_df_sheet(wb, "07_Lift_Tables", combined)

    def _write_validation_sheet(self, wb: Workbook, validation_report: Any) -> None:
        """Write the 08_Validation sheet from a ValidationReport."""
        ws = wb.create_sheet("08_Validation")
        ws.sheet_properties.tabColor = "E74C3C"

        ws["A1"].comment = Comment(
            "Validation check results: data quality and model quality.", "ExcelReporter"
        )

        # Build DataFrame
        if hasattr(validation_report, "to_dataframe"):
            df = validation_report.to_dataframe()
        elif isinstance(validation_report, pd.DataFrame):
            df = validation_report
        else:
            ws["A1"] = "No validation data"
            return

        if df.empty:
            ws["A1"] = "No validation checks ran"
            return

        headers = list(df.columns)
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        status_idx = headers.index("Status") + 1 if "Status" in headers else None

        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER
                cell.value = _safe_value(value)

            # Row-level conditional formatting based on Status
            if status_idx:
                status_val = str(ws.cell(row=row_idx, column=status_idx).value).upper()
                fill_map = {"PASS": PASS_FILL, "WARNING": WARN_FILL, "FAIL": FAIL_FILL}
                fill = fill_map.get(status_val)
                if fill:
                    for ci in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=ci).fill = fill

        _autofit_columns(ws, df)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _write_score_distribution_sheet(
        self, wb: Workbook, score_dist: Dict[str, Any]
    ) -> None:
        """Write 09_Score_Distribution sheet.

        Expects score_dist keys:
            histograms: dict of period -> list of dicts with bin_edge, count
            psi_values: list of dicts with Period, PSI
        """
        ws = wb.create_sheet("09_Score_Distribution")
        ws["A1"].comment = Comment(
            "Score histogram per period and score PSI values.", "ExcelReporter"
        )

        row = 1

        # Histograms
        histograms = score_dist.get("histograms", {})
        if histograms:
            ws.cell(row=row, column=1, value="Score Histograms").font = Font(
                bold=True, size=12, color="2F5496"
            )
            row += 1
            for period, bins in histograms.items():
                if isinstance(bins, pd.DataFrame):
                    bins_df = bins
                elif isinstance(bins, list):
                    bins_df = pd.DataFrame(bins)
                else:
                    continue
                ws.cell(row=row, column=1, value=f"Period: {period}").font = Font(bold=True)
                row += 1
                for col_idx, col_name in enumerate(bins_df.columns, 1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER
                row += 1
                for _, rd in bins_df.iterrows():
                    for col_idx, value in enumerate(rd, 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = _safe_value(value)
                        cell.border = THIN_BORDER
                    row += 1
                row += 1  # blank row between periods

        # PSI values
        psi_values = score_dist.get("psi_values")
        if psi_values is not None:
            if isinstance(psi_values, list):
                psi_df = pd.DataFrame(psi_values)
            elif isinstance(psi_values, pd.DataFrame):
                psi_df = psi_values
            else:
                psi_df = pd.DataFrame()

            if not psi_df.empty:
                ws.cell(row=row, column=1, value="Score PSI").font = Font(
                    bold=True, size=12, color="2F5496"
                )
                row += 1
                for col_idx, col_name in enumerate(psi_df.columns, 1):
                    cell = ws.cell(row=row, column=col_idx, value=col_name)
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.border = THIN_BORDER
                row += 1
                for _, rd in psi_df.iterrows():
                    for col_idx, value in enumerate(rd, 1):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.value = _safe_value(value)
                        cell.border = THIN_BORDER
                    row += 1

    def _write_config_sheet(
        self, wb: Workbook, config: Any
    ) -> None:
        """Write 10_Config sheet as section/parameter/value table."""
        ws = wb.create_sheet("10_Config")
        ws["A1"].comment = Comment(
            "Full pipeline configuration used for this run.", "ExcelReporter"
        )

        headers = ["Section", "Parameter", "Value"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Flatten config to rows
        rows = _flatten_config(config)

        for row_idx, (section, param, value) in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=section).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=param).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=str(value)).border = THIN_BORDER

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 50
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions


# ==========================================================================
# Helpers
# ==========================================================================


def _safe_value(value: Any) -> Any:
    """Convert numpy/pandas types to native Python for openpyxl."""
    if pd.isna(value):
        return None
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def _autofit_columns(ws: Any, df: pd.DataFrame) -> None:
    """Approximate auto-fit for column widths."""
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 102)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max_len + 3, 40)


def _flatten_config(config: Any) -> List[tuple]:
    """Flatten a PipelineConfig (or dict) into (section, parameter, value) rows."""
    rows: List[tuple] = []

    if hasattr(config, "model_dump"):
        # Pydantic v2
        data = config.model_dump()
    elif hasattr(config, "dict"):
        # Pydantic v1
        data = config.dict()
    elif isinstance(config, dict):
        data = config
    else:
        rows.append(("config", "raw", str(config)))
        return rows

    def _recurse(d: dict, section: str) -> None:
        for key, val in d.items():
            if isinstance(val, dict):
                _recurse(val, f"{section}.{key}" if section else key)
            elif isinstance(val, list):
                rows.append((section, key, str(val)))
            else:
                rows.append((section, key, val))

    _recurse(data, "")
    return rows
