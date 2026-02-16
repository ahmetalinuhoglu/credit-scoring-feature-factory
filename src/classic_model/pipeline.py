"""
Classic Credit Risk Model Pipeline

Orchestrates the WoE + Logistic Regression + Scorecard model development:

1. Load and split data (Train / Test / OOT)
2. WoE binning (fit on train)
3. IV-based feature selection
4. PSI stability check on WoE-encoded values
5. Correlation elimination on WoE-encoded features
6. Logistic Regression fit on WoE-encoded features (with StandardScaler)
7. Scorecard generation
8. Model evaluation (via ClassicModelAdapter + evaluator.py)
9. Excel report generation

Differs from the XGBoost pipeline in several ways:
- Features are WoE-encoded (monotonic binning) rather than used raw.
- Model is an interpretable LogReg, not a boosted ensemble.
- A point-based scorecard is produced alongside standard evaluation.
"""

from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from pathlib import Path
import logging

import numpy as np
import pandas as pd
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler

from src.config.schema import ClassicPipelineConfig
from src.model_development.data_loader import load_and_split, DataSets
from src.model_development.evaluator import (
    evaluate_model_quarterly,
    bootstrap_auc_ci,
    compute_score_psi,
)
from src.features.woe_transformer import WoETransformer
from src.classic_model.model_adapter import ClassicModelAdapter
from src.classic_model.scorecard import ScorecardGenerator


logger = logging.getLogger(__name__)


class ClassicModelPipeline:
    """End-to-end classic credit scoring pipeline.

    Parameters
    ----------
    input_path : str
        Path to the features parquet / csv file.
    train_end_date : str
        Cutoff date (YYYY-MM-DD). Data up to this date is train+test;
        data after is OOT.
    output_dir : str
        Base directory for outputs (Excel, model, etc.).
    config : ClassicPipelineConfig, optional
        Typed configuration object.  If ``None``, a default is created.
    """

    def __init__(
        self,
        input_path: str,
        train_end_date: str,
        output_dir: str = "outputs/classic_model",
        config: Optional[ClassicPipelineConfig] = None,
    ):
        self.input_path = input_path
        self.train_end_date = train_end_date
        self.output_dir = output_dir
        self.config = config or ClassicPipelineConfig()

        # Runtime artefacts (populated during run)
        self.datasets: Optional[DataSets] = None
        self.woe_transformer: Optional[WoETransformer] = None
        self.model: Optional[LogisticRegression] = None
        self.scaler: Optional[StandardScaler] = None
        self.adapter: Optional[ClassicModelAdapter] = None
        self.scorecard_gen: Optional[ScorecardGenerator] = None

    # ------------------------------------------------------------------
    # Public entry point
    # ------------------------------------------------------------------

    def run(self) -> Dict[str, Any]:
        """Execute the full pipeline and return a results dictionary."""
        start_time = datetime.now()
        run_id = start_time.strftime("%Y%m%d_%H%M%S")
        run_dir = Path(self.output_dir) / run_id
        run_dir.mkdir(parents=True, exist_ok=True)

        # Reproducibility
        seed = self.config.reproducibility.global_seed
        np.random.seed(seed)

        logger.info("=" * 60)
        logger.info("Classic Model Pipeline  --  run_id=%s", run_id)
        logger.info("=" * 60)

        # 1. Load and split
        logger.info("Step 1/9: Loading and splitting data")
        self.datasets = self._load_data()
        features = list(self.datasets.feature_columns)
        target = self.datasets.target_column

        # 2. WoE binning
        logger.info("Step 2/9: WoE binning on %d features", len(features))
        self.woe_transformer, woe_train, woe_test, woe_oot = self._fit_woe(features)

        # 3. IV-based selection
        logger.info("Step 3/9: IV-based feature selection")
        selected_features, iv_summary = self._iv_selection()

        # WoE column names for selected features
        woe_selected = [f"{f}_woe" for f in selected_features]

        # 4. PSI check on WoE-encoded values
        logger.info("Step 4/9: PSI check on WoE-encoded features")
        woe_selected, psi_details = self._psi_check(woe_train, woe_selected)

        # Map back to original feature names after PSI
        selected_features = [c.replace("_woe", "") for c in woe_selected]

        # 5. Correlation elimination on WoE-encoded features
        logger.info("Step 5/9: Correlation elimination")
        woe_selected, corr_details = self._correlation_elimination(
            woe_train, woe_selected, iv_summary,
        )
        selected_features = [c.replace("_woe", "") for c in woe_selected]

        if len(woe_selected) == 0:
            logger.error("No features survived filtering. Aborting.")
            return {
                "status": "failed",
                "reason": "no_features",
                "run_id": run_id,
            }

        # 6. LogReg fit
        logger.info(
            "Step 6/9: Fitting Logistic Regression on %d WoE features",
            len(woe_selected),
        )
        self.model, self.scaler = self._fit_logistic(
            woe_train, woe_selected, target,
        )
        self.adapter = ClassicModelAdapter(
            self.model, self.scaler, woe_selected,
        )

        # 7. Scorecard generation
        logger.info("Step 7/9: Generating scorecard")
        scorecard_df = self._generate_scorecard(selected_features)

        # 8. Evaluation
        logger.info("Step 8/9: Evaluating model")
        eval_results = self._evaluate(
            woe_train, woe_test, woe_oot, woe_selected, target,
        )
        performance_df = eval_results["performance_df"]
        lift_tables = eval_results["lift_tables"]
        importance_df = eval_results["importance_df"]
        bootstrap_df = eval_results.get("bootstrap_df")
        score_psi_df = eval_results.get("score_psi_df")

        # Compute scorecard scores for reporting
        train_scores = self.scorecard_gen.score(
            self.datasets.train, self.woe_transformer, selected_features,
        )
        logger.info(
            "Scorecard scores: min=%d, max=%d, mean=%.1f",
            train_scores.min(), train_scores.max(), train_scores.mean(),
        )

        # 9. Excel report
        logger.info("Step 9/9: Generating Excel report")
        excel_path = str(run_dir / f"classic_model_{run_id}.xlsx")
        self._generate_excel(
            excel_path=excel_path,
            run_id=run_id,
            selected_features=selected_features,
            woe_selected=woe_selected,
            iv_summary=iv_summary,
            psi_details=psi_details,
            corr_details=corr_details,
            scorecard_df=scorecard_df,
            performance_df=performance_df,
            lift_tables=lift_tables,
            importance_df=importance_df,
            bootstrap_df=bootstrap_df,
            score_psi_df=score_psi_df,
        )

        # Save model artefacts
        model_path = None
        if self.config.output.save_model:
            model_path = str(run_dir / "classic_model.joblib")
            self._save_model(model_path)

        # Save WoE binning config
        woe_path = str(run_dir / "woe_binning.json")
        try:
            self.woe_transformer.export_binning(woe_path)
        except Exception as exc:
            logger.warning("Could not save WoE binning: %s", exc)

        duration = (datetime.now() - start_time).total_seconds()
        logger.info("Pipeline completed in %.1f seconds", duration)

        return {
            "status": "success",
            "run_id": run_id,
            "run_dir": str(run_dir),
            "selected_features": selected_features,
            "n_selected": len(selected_features),
            "excel_path": excel_path,
            "model_path": model_path,
            "woe_path": woe_path,
            "duration_seconds": round(duration, 1),
        }

    # ==================================================================
    # Step implementations
    # ==================================================================

    def _load_data(self) -> DataSets:
        """Step 1: Load features and split into Train/Test/OOT."""
        cfg = self.config
        return load_and_split(
            input_path=self.input_path,
            train_end_date=self.train_end_date,
            target_column=cfg.data.target_column,
            date_column=cfg.data.date_column,
            id_columns=list(cfg.data.id_columns),
            meta_columns=list(cfg.data.exclude_columns),
            test_size=cfg.splitting.test_size,
            stratify=cfg.splitting.stratify,
            random_state=cfg.reproducibility.global_seed,
        )

    # ------------------------------------------------------------------

    def _fit_woe(
        self,
        features: List[str],
    ) -> Tuple[WoETransformer, pd.DataFrame, pd.DataFrame, Dict[str, pd.DataFrame]]:
        """Step 2: Fit WoE binning on train; transform all sets."""
        cfg = self.config
        target = cfg.data.target_column

        # The WoETransformer expects a nested config dict
        woe_config = {
            "model": {
                "logistic_regression": {
                    "woe_binning": {
                        "n_bins": cfg.woe.n_bins,
                        "min_bin_size": cfg.woe.min_bin_size,
                        "monotonic": cfg.woe.monotonic,
                        "missing_bin": cfg.woe.missing_bin,
                    }
                }
            }
        }

        transformer = WoETransformer(config=woe_config, name="ClassicWoE")

        # Fit on train
        transformer.fit(self.datasets.train, features, target_column=target)
        logger.info(
            "WoE binning fitted on %d features",
            len(transformer.fitted_features),
        )

        # Transform all sets
        woe_train = transformer.transform(self.datasets.train)
        woe_test = transformer.transform(self.datasets.test)
        woe_oot: Dict[str, pd.DataFrame] = {}
        for label, qdf in self.datasets.oot_quarters.items():
            woe_oot[label] = transformer.transform(qdf)

        return transformer, woe_train, woe_test, woe_oot

    # ------------------------------------------------------------------

    def _iv_selection(self) -> Tuple[List[str], Dict[str, float]]:
        """Step 3: Select features by IV range [min_iv, max_iv]."""
        cfg = self.config
        iv_summary = self.woe_transformer.get_iv_summary()

        selected = []
        for feat, iv in iv_summary.items():
            if cfg.woe.min_iv <= iv <= cfg.woe.max_iv:
                selected.append(feat)
            else:
                reason = "too low" if iv < cfg.woe.min_iv else "suspicious"
                logger.debug(
                    "IV filter: dropping '%s' (IV=%.4f, %s)", feat, iv, reason,
                )

        # Sort by IV descending for downstream steps
        selected.sort(key=lambda f: iv_summary.get(f, 0), reverse=True)

        logger.info(
            "IV selection: %d -> %d features (min_iv=%.2f, max_iv=%.2f)",
            len(iv_summary), len(selected), cfg.woe.min_iv, cfg.woe.max_iv,
        )
        return selected, iv_summary

    # ------------------------------------------------------------------

    def _psi_check(
        self,
        woe_train: pd.DataFrame,
        woe_features: List[str],
    ) -> Tuple[List[str], pd.DataFrame]:
        """Step 4: PSI check on WoE-encoded values within training data.

        Splits training data into first-half and second-half by date and
        computes PSI per feature.
        """
        cfg = self.config
        psi_threshold = 0.25  # default
        date_col = cfg.data.date_column

        train_dates = woe_train[date_col]
        median_date = train_dates.median()
        mask_first = train_dates <= median_date
        mask_second = train_dates > median_date

        first_half = woe_train.loc[mask_first]
        second_half = woe_train.loc[mask_second]

        psi_rows = []
        kept = []

        for feat in woe_features:
            try:
                psi_val = self._calculate_psi(
                    first_half[feat].values,
                    second_half[feat].values,
                )
            except Exception as exc:
                logger.warning("PSI calc failed for '%s': %s", feat, exc)
                psi_val = None

            status = "N/A"
            if psi_val is not None:
                if psi_val < 0.10:
                    status = "Stable"
                elif psi_val < psi_threshold:
                    status = "Moderate"
                else:
                    status = "Unstable"

            psi_rows.append({
                "Feature": feat,
                "PSI": round(psi_val, 4) if psi_val is not None else None,
                "Status": status,
            })

            if psi_val is None or psi_val < psi_threshold:
                kept.append(feat)
            else:
                logger.info(
                    "PSI filter: dropping '%s' (PSI=%.4f)", feat, psi_val,
                )

        details_df = pd.DataFrame(psi_rows)
        n_dropped = len(woe_features) - len(kept)
        logger.info(
            "PSI check: %d -> %d features (%d unstable)",
            len(woe_features), len(kept), n_dropped,
        )
        return kept, details_df

    # ------------------------------------------------------------------

    def _correlation_elimination(
        self,
        woe_train: pd.DataFrame,
        woe_features: List[str],
        iv_summary: Dict[str, float],
    ) -> Tuple[List[str], pd.DataFrame]:
        """Step 5: Greedy correlation removal (keep higher-IV feature)."""
        threshold = 0.80  # default

        if len(woe_features) < 2:
            return woe_features, pd.DataFrame()

        corr_matrix = woe_train[woe_features].corr(method="pearson").abs()

        # Build a set of features to drop
        to_drop = set()
        pairs = []

        for i in range(len(woe_features)):
            for j in range(i + 1, len(woe_features)):
                feat_a = woe_features[i]
                feat_b = woe_features[j]
                corr_val = corr_matrix.iloc[i, j]

                if corr_val >= threshold:
                    # Determine which to drop (lower IV)
                    orig_a = feat_a.replace("_woe", "")
                    orig_b = feat_b.replace("_woe", "")
                    iv_a = iv_summary.get(orig_a, 0)
                    iv_b = iv_summary.get(orig_b, 0)

                    if iv_a >= iv_b:
                        drop_feat = feat_b
                        keep_feat = feat_a
                    else:
                        drop_feat = feat_a
                        keep_feat = feat_b

                    to_drop.add(drop_feat)
                    pairs.append({
                        "Feature_A": feat_a,
                        "Feature_B": feat_b,
                        "Correlation": round(corr_val, 4),
                        "Dropped": drop_feat,
                        "Kept": keep_feat,
                    })

        kept = [f for f in woe_features if f not in to_drop]
        pairs_df = pd.DataFrame(pairs) if pairs else pd.DataFrame()

        logger.info(
            "Correlation elimination: %d -> %d features (%d dropped, threshold=%.2f)",
            len(woe_features), len(kept), len(to_drop), threshold,
        )
        return kept, pairs_df

    # ------------------------------------------------------------------

    def _fit_logistic(
        self,
        woe_train: pd.DataFrame,
        woe_features: List[str],
        target: str,
    ) -> Tuple[LogisticRegression, StandardScaler]:
        """Step 6: Fit LogReg on WoE-encoded features (with scaling)."""
        cfg = self.config

        X = woe_train[woe_features].values
        y = woe_train[target].values

        # Scale
        scaler = StandardScaler()
        X_scaled = scaler.fit_transform(X)

        # Resolve class_weight
        class_weight = cfg.logistic.class_weight
        if class_weight == "none" or class_weight is None:
            class_weight = None

        # Resolve penalty
        penalty = cfg.logistic.penalty
        if penalty == "none":
            penalty = None

        model = LogisticRegression(
            solver=cfg.logistic.solver,
            penalty=penalty,
            C=cfg.logistic.C,
            max_iter=cfg.logistic.max_iter,
            class_weight=class_weight,
            random_state=cfg.reproducibility.global_seed,
            n_jobs=cfg.reproducibility.n_jobs,
        )
        model.fit(X_scaled, y)

        # Log coefficients
        for name, coef in zip(woe_features, model.coef_[0]):
            logger.info("  Coefficient  %s = %.4f", name, coef)
        logger.info("  Intercept = %.4f", model.intercept_[0])

        return model, scaler

    # ------------------------------------------------------------------

    def _generate_scorecard(
        self,
        selected_features: List[str],
    ) -> pd.DataFrame:
        """Step 7: Build the point-based scorecard."""
        cfg = self.config
        self.scorecard_gen = ScorecardGenerator(
            target_score=cfg.scorecard.target_score,
            target_odds=cfg.scorecard.target_odds,
            pdo=cfg.scorecard.pdo,
        )
        scorecard_df = self.scorecard_gen.generate(
            model=self.model,
            scaler=self.scaler,
            woe_transformer=self.woe_transformer,
            feature_names=selected_features,
        )
        logger.info("Scorecard: %d rows across %d features", len(scorecard_df), len(selected_features))
        return scorecard_df

    # ------------------------------------------------------------------

    def _evaluate(
        self,
        woe_train: pd.DataFrame,
        woe_test: pd.DataFrame,
        woe_oot: Dict[str, pd.DataFrame],
        woe_features: List[str],
        target: str,
    ) -> Dict[str, Any]:
        """Step 8: Evaluate model using the shared evaluator."""
        # evaluate_model_quarterly internally calls model.predict_proba(X)[:, 1]
        # and model.get_booster() for importance.  For the classic adapter we
        # provide coefficient-based importance separately.

        # Build evaluation DataFrames containing only WoE columns
        # (adapter.predict_proba will select the right columns)
        performance_df, lift_tables, _ = evaluate_model_quarterly(
            model=self.adapter,
            selected_features=woe_features,
            train_df=woe_train,
            test_df=woe_test,
            oot_quarters=woe_oot,
            target_column=target,
        )

        # Coefficient-based feature importance
        abs_coefs = np.abs(self.model.coef_[0])
        total = abs_coefs.sum()
        if total > 0:
            importances = abs_coefs / total
        else:
            importances = abs_coefs

        importance_df = pd.DataFrame({
            "Feature": woe_features,
            "Coefficient": self.model.coef_[0],
            "Importance": importances,
        }).sort_values("Importance", ascending=False)
        importance_df["Rank"] = range(1, len(importance_df) + 1)
        importance_df["Cumulative_Importance"] = importance_df["Importance"].cumsum()
        importance_df = importance_df.reset_index(drop=True)

        result: Dict[str, Any] = {
            "performance_df": performance_df,
            "lift_tables": lift_tables,
            "importance_df": importance_df,
        }

        # Bootstrap CI
        cfg = self.config
        if cfg.evaluation.bootstrap.enabled:
            try:
                periods = [("Train", woe_train), ("Test", woe_test)]
                for label in sorted(woe_oot.keys()):
                    periods.append((f"OOT_{label}", woe_oot[label]))

                bootstrap_df = bootstrap_auc_ci(
                    model=self.adapter,
                    selected_features=woe_features,
                    datasets=periods,
                    target_column=target,
                    n_iterations=cfg.evaluation.bootstrap.n_iterations,
                    confidence_level=cfg.evaluation.bootstrap.confidence_level,
                    n_jobs=cfg.reproducibility.n_jobs,
                )
                result["bootstrap_df"] = bootstrap_df

                # Merge CI into performance_df
                if bootstrap_df is not None and not bootstrap_df.empty:
                    ci_cols = bootstrap_df[["Period", "CI_Lower", "CI_Upper"]].copy()
                    performance_df = performance_df.merge(ci_cols, on="Period", how="left")
                    result["performance_df"] = performance_df
            except Exception as exc:
                logger.warning("Bootstrap CI failed: %s", exc)

        # Score PSI
        if cfg.evaluation.calculate_score_psi:
            try:
                train_probs = self.adapter.predict_proba(woe_train[woe_features])[:, 1]
                oot_scores = {}
                for label in sorted(woe_oot.keys()):
                    qdf = woe_oot[label]
                    oot_scores[f"OOT_{label}"] = (
                        self.adapter.predict_proba(qdf[woe_features])[:, 1]
                    )
                score_psi_df = compute_score_psi(train_probs, oot_scores)
                result["score_psi_df"] = score_psi_df
            except Exception as exc:
                logger.warning("Score PSI failed: %s", exc)

        return result

    # ------------------------------------------------------------------

    def _generate_excel(
        self,
        excel_path: str,
        run_id: str,
        selected_features: List[str],
        woe_selected: List[str],
        iv_summary: Dict[str, float],
        psi_details: pd.DataFrame,
        corr_details: pd.DataFrame,
        scorecard_df: pd.DataFrame,
        performance_df: pd.DataFrame,
        lift_tables: Dict[str, pd.DataFrame],
        importance_df: pd.DataFrame,
        bootstrap_df: Optional[pd.DataFrame] = None,
        score_psi_df: Optional[pd.DataFrame] = None,
    ) -> str:
        """Step 9: Write Excel report using openpyxl."""
        import numpy as np  # noqa: F811  (needed for openpyxl np.integer check)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows

        wb = Workbook()

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font_white = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        def _write_df(ws, df, start_row=1):
            """Write DataFrame to worksheet with formatted headers."""
            for r_idx, row in enumerate(
                dataframe_to_rows(df, index=False, header=True), start_row,
            ):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = thin_border
                    if r_idx == start_row:
                        cell.font = header_font_white
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = "01_Summary"
        summary_data = self._build_summary(
            run_id, selected_features, iv_summary, performance_df,
        )
        for r_idx, (key, val) in enumerate(summary_data.items(), 1):
            ws.cell(row=r_idx, column=1, value=str(key)).font = header_font
            ws.cell(row=r_idx, column=2, value=str(val))
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 60

        # --- Sheet 2: IV Summary ---
        ws_iv = wb.create_sheet("02_IV_Summary")
        iv_df = pd.DataFrame([
            {
                "Feature": f,
                "IV": round(iv, 4),
                "Category": self.woe_transformer.get_iv_category(iv),
                "Selected": f in selected_features,
            }
            for f, iv in sorted(iv_summary.items(), key=lambda x: -x[1])
        ])
        _write_df(ws_iv, iv_df)

        # --- Sheet 3: PSI Details ---
        ws_psi = wb.create_sheet("03_PSI_Check")
        if not psi_details.empty:
            _write_df(ws_psi, psi_details)
        else:
            ws_psi.cell(row=1, column=1, value="No PSI data")

        # --- Sheet 4: Correlation ---
        ws_corr = wb.create_sheet("04_Correlation")
        if not corr_details.empty:
            _write_df(ws_corr, corr_details)
        else:
            ws_corr.cell(row=1, column=1, value="No correlated pairs above threshold")

        # --- Sheet 5: Coefficients ---
        ws_coef = wb.create_sheet("05_Coefficients")
        _write_df(ws_coef, importance_df)

        # --- Sheet 6: Scorecard ---
        ws_sc = wb.create_sheet("06_Scorecard")
        _write_df(ws_sc, scorecard_df)

        # --- Sheet 7: Performance ---
        ws_perf = wb.create_sheet("07_Performance")
        _write_df(ws_perf, performance_df)

        # --- Sheet 8: Lift tables ---
        ws_lift = wb.create_sheet("08_Lift_Tables")
        current_row = 1
        for period_name, lt in lift_tables.items():
            ws_lift.cell(row=current_row, column=1, value=period_name).font = header_font
            current_row += 1
            _write_df(ws_lift, lt, start_row=current_row)
            current_row += len(lt) + 2

        # --- Sheet 9: Bootstrap CI ---
        if bootstrap_df is not None and not bootstrap_df.empty:
            ws_boot = wb.create_sheet("09_Bootstrap_CI")
            _write_df(ws_boot, bootstrap_df)

        # --- Sheet 10: Score PSI ---
        if score_psi_df is not None and not score_psi_df.empty:
            ws_spsi = wb.create_sheet("10_Score_PSI")
            _write_df(ws_spsi, score_psi_df)

        # --- Sheet 11: WoE Bins ---
        ws_woe = wb.create_sheet("11_WoE_Bins")
        woe_rows = []
        for feat in selected_features:
            bins = self.woe_transformer.get_woe_table(feat)
            if bins is None:
                continue
            for b in bins:
                bad_rate = b.bad_count / b.count if b.count > 0 else 0
                woe_rows.append({
                    "Feature": feat,
                    "Bin_ID": b.bin_id,
                    "Lower": b.lower_bound,
                    "Upper": b.upper_bound,
                    "Count": b.count,
                    "Good": b.good_count,
                    "Bad": b.bad_count,
                    "Bad_Rate": round(bad_rate, 4),
                    "WoE": round(b.woe, 4),
                    "IV_Contribution": round(b.iv_contribution, 6),
                })
        if woe_rows:
            _write_df(ws_woe, pd.DataFrame(woe_rows))
        else:
            ws_woe.cell(row=1, column=1, value="No WoE bin data")

        # Save
        wb.save(excel_path)
        logger.info("Excel report saved: %s", excel_path)
        return excel_path

    # ------------------------------------------------------------------

    def _build_summary(
        self,
        run_id: str,
        selected_features: List[str],
        iv_summary: Dict[str, float],
        performance_df: pd.DataFrame,
    ) -> Dict[str, str]:
        """Build the summary dictionary for the first Excel sheet."""
        cfg = self.config
        ds = self.datasets
        target = cfg.data.target_column
        train_dates = ds.train[cfg.data.date_column]

        summary = {
            "Run Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Run ID": run_id,
            "Pipeline Type": "Classic (WoE + LogReg + Scorecard)",
            "Input File": self.input_path,
            "Train End Date": self.train_end_date,
            "Train Period": (
                f"{train_dates.min().strftime('%Y-%m-%d')} to "
                f"{train_dates.max().strftime('%Y-%m-%d')}"
            ),
            "OOT Periods": ", ".join(ds.oot_labels) if ds.oot_labels else "None",
            "Train Rows": str(len(ds.train)),
            "Test Rows": str(len(ds.test)),
            "Train Bad Rate": f"{ds.train[target].mean():.2%}",
            "Test Bad Rate": f"{ds.test[target].mean():.2%}",
            "": "",
            "Total Features": str(len(ds.feature_columns)),
            "Fitted by WoE": str(len(iv_summary)),
            "After IV Filter": str(len(selected_features)),
            "Final Selected": str(len(selected_features)),
            " ": "",
            "WoE n_bins": str(cfg.woe.n_bins),
            "WoE monotonic": str(cfg.woe.monotonic),
            "IV Range": f"[{cfg.woe.min_iv}, {cfg.woe.max_iv}]",
            "LogReg Solver": cfg.logistic.solver,
            "LogReg Penalty": cfg.logistic.penalty,
            "LogReg C": str(cfg.logistic.C),
            "Scorecard Target Score": str(cfg.scorecard.target_score),
            "Scorecard Target Odds": str(cfg.scorecard.target_odds),
            "Scorecard PDO": str(cfg.scorecard.pdo),
        }

        # Add performance metrics
        summary["  "] = ""
        for _, row in performance_df.iterrows():
            period = row["Period"]
            summary[f"AUC {period}"] = str(row.get("AUC", ""))
            summary[f"Gini {period}"] = str(row.get("Gini", ""))
            summary[f"KS {period}"] = str(row.get("KS", ""))

        return summary

    # ------------------------------------------------------------------

    def _save_model(self, path: str) -> None:
        """Save model + scaler + metadata to disk."""
        import joblib
        artefact = {
            "model": self.model,
            "scaler": self.scaler,
            "feature_names": self.adapter.feature_names if self.adapter else [],
            "config": self.config.model_dump(),
        }
        joblib.dump(artefact, path)
        logger.info("Model artefact saved: %s", path)

    # ------------------------------------------------------------------
    # PSI helper
    # ------------------------------------------------------------------

    @staticmethod
    def _calculate_psi(
        expected: np.ndarray,
        actual: np.ndarray,
        n_bins: int = 10,
    ) -> Optional[float]:
        """Calculate Population Stability Index between two distributions."""
        try:
            expected = expected[~np.isnan(expected)]
            actual = actual[~np.isnan(actual)]

            if len(expected) < 10 or len(actual) < 10:
                return None

            # Build bins from expected distribution
            try:
                _, bin_edges = pd.qcut(
                    expected, q=n_bins, retbins=True, duplicates="drop",
                )
            except ValueError:
                n = min(5, len(np.unique(expected)))
                if n < 2:
                    return None
                _, bin_edges = pd.qcut(
                    expected, q=n, retbins=True, duplicates="drop",
                )

            bin_edges[0] = -np.inf
            bin_edges[-1] = np.inf

            expected_bins = pd.cut(expected, bins=bin_edges)
            actual_bins = pd.cut(actual, bins=bin_edges)

            expected_pct = (
                pd.Series(expected_bins)
                .value_counts(normalize=True)
                .sort_index()
            )
            actual_pct = (
                pd.Series(actual_bins)
                .value_counts(normalize=True)
                .sort_index()
            )

            all_bins = expected_pct.index.union(actual_pct.index)
            expected_pct = expected_pct.reindex(all_bins, fill_value=1e-4)
            actual_pct = actual_pct.reindex(all_bins, fill_value=1e-4)

            epsilon = 1e-4
            expected_pct = expected_pct.clip(lower=epsilon)
            actual_pct = actual_pct.clip(lower=epsilon)

            psi = float(
                ((actual_pct - expected_pct) * np.log(actual_pct / expected_pct)).sum()
            )
            return psi if np.isfinite(psi) else None
        except Exception:
            return None
