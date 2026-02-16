"""
Excel Reporter

Generates a single Excel workbook with all elimination and evaluation details.
"""

from typing import Any, Dict, List, Optional
from datetime import datetime
from pathlib import Path
import logging

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils.dataframe import dataframe_to_rows


logger = logging.getLogger(__name__)


# Styles
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
KEPT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ELIM_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
OPTIMAL_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Centralized sheet name mapping (clean sequential numbering)
SHEET_NAMES = {
    'summary': '00_Summary',
    'constant': '01_Constant',
    'missing': '02_Missing',
    'iv': '03_IV_Analysis',
    'psi': '04_PSI_Stability',
    'correlation': '05_Correlation',
    'corr_pairs': '06_Corr_Pairs',
    'temporal_filter': '07_Temporal_Filter',
    'selection': '08_Selection',
    'vif': '09_VIF',
    'tuning': '10_Tuning',
    'tuning_best': '10a_Tuning_Best',
    'performance': '11_Performance',
    'quarterly_performance': '11a_Quarterly_Perf',
    'variable_quarterly': '11b_Var_Quarterly_AUC',
    'lift_tables': '12_Lift_Tables',
    'importance': '13_Importance',
    'confusion_matrix': '14_Confusion_Matrix',
    'score_psi': '15_Score_PSI',
    'bootstrap_ci': '16_Bootstrap_CI',
    'shap': '17_SHAP',
    'calibration': '18_Calibration',
    'subseg_perf': '19_Subseg_Perf',
    'subseg_confusion': '20_Subseg_Confusion',
    'validation': '21_Validation',
}


def generate_report(
    output_path: str,
    summary: Dict[str, Any],
    elimination_results: List[Any],
    corr_pairs_df: Optional[pd.DataFrame],
    selection_df: pd.DataFrame,
    performance_df: pd.DataFrame,
    lift_tables: Dict[str, pd.DataFrame],
    importance_df: pd.DataFrame,
    vif_df: Optional[pd.DataFrame] = None,
    tuning_df: Optional[pd.DataFrame] = None,
    tuning_best_params: Optional[Dict] = None,
    chart_path: Optional[str] = None,
    score_psi_df: Optional[pd.DataFrame] = None,
    bootstrap_df: Optional[pd.DataFrame] = None,
    shap_summary_df: Optional[pd.DataFrame] = None,
    shap_plot_path: Optional[str] = None,
    calibration_dict: Optional[Dict[str, Any]] = None,
    validation_report_df: Optional[pd.DataFrame] = None,
    # NEW parameters
    quarterly_perf_df: Optional[pd.DataFrame] = None,
    variable_quarterly_df: Optional[pd.DataFrame] = None,
    confusion_matrix_df: Optional[pd.DataFrame] = None,
    subsegment_perf: Optional[Dict[str, pd.DataFrame]] = None,
    subsegment_confusion: Optional[Dict[str, pd.DataFrame]] = None,
) -> str:
    """
    Generate the full Excel report.

    Args:
        output_path: Path for the output Excel file.
        summary: Dict of summary key-value pairs.
        elimination_results: List of EliminationResult objects.
        corr_pairs_df: Correlation pairs details DataFrame.
        selection_df: Sequential selection details DataFrame.
        performance_df: Quarterly performance DataFrame.
        lift_tables: Dict of period -> lift table DataFrame.
        importance_df: Feature importance DataFrame.
        vif_df: VIF elimination details DataFrame (optional).
        tuning_df: Optuna trial history DataFrame (optional).
        tuning_best_params: Best hyperparameters dict (optional).
        chart_path: Path to selection performance chart PNG (optional).
        score_psi_df: Score PSI DataFrame (optional).
        bootstrap_df: Bootstrap CI DataFrame (optional).
        shap_summary_df: SHAP summary DataFrame (optional).
        shap_plot_path: Path to SHAP plot PNG (optional).
        calibration_dict: Calibration results dict (optional).
        validation_report_df: Validation report DataFrame (optional).
        quarterly_perf_df: Quarterly performance DataFrame (optional).
        variable_quarterly_df: Per-variable quarterly AUC DataFrame (optional).
        confusion_matrix_df: Confusion matrix DataFrame (optional).
        subsegment_perf: Dict of subsegment -> performance DataFrame (optional).
        subsegment_confusion: Dict of subsegment -> confusion matrix DataFrame (optional).

    Returns:
        Path to the generated Excel file.
    """
    wb = Workbook()

    # 00_Summary
    _write_summary_sheet(wb, summary)

    # Elimination sheets (01-05, 07_Temporal_Filter, 09_VIF via step_name)
    for result in elimination_results:
        _write_df_sheet(wb, result.step_name, result.details_df)

    # 06_Corr_Pairs
    if corr_pairs_df is not None and len(corr_pairs_df) > 0:
        _write_df_sheet(wb, SHEET_NAMES['corr_pairs'], corr_pairs_df)

    # 08_Selection
    _write_selection_sheet(wb, SHEET_NAMES['selection'], selection_df, chart_path)

    # 09_VIF (if not already added via elimination_results)
    vif_already_added = any(
        r.step_name == SHEET_NAMES['vif'] for r in elimination_results
    )
    if vif_df is not None and not vif_already_added:
        _write_df_sheet(wb, SHEET_NAMES['vif'], vif_df)

    # 10_Tuning
    if tuning_df is not None:
        _write_df_sheet(wb, SHEET_NAMES['tuning'], tuning_df)

    # 10a_Tuning_Best
    if tuning_best_params is not None:
        _write_best_params_sheet(wb, tuning_best_params)

    # 11_Performance
    _write_df_sheet(wb, SHEET_NAMES['performance'], performance_df)

    # 11a_Quarterly_Perf
    if quarterly_perf_df is not None and len(quarterly_perf_df) > 0:
        _write_df_sheet(wb, SHEET_NAMES['quarterly_performance'], quarterly_perf_df)

    # 11b_Var_Quarterly_AUC
    if variable_quarterly_df is not None and len(variable_quarterly_df) > 0:
        _write_df_sheet(wb, SHEET_NAMES['variable_quarterly'], variable_quarterly_df)

    # 12_Lift_Tables (one sheet per period, or combined)
    _write_lift_sheets(wb, lift_tables)

    # 13_Importance
    _write_df_sheet(wb, SHEET_NAMES['importance'], importance_df)

    # 14_Confusion_Matrix
    if confusion_matrix_df is not None and len(confusion_matrix_df) > 0:
        _write_df_sheet(wb, SHEET_NAMES['confusion_matrix'], confusion_matrix_df)

    # 15_Score_PSI
    if score_psi_df is not None and len(score_psi_df) > 0:
        _write_df_sheet(wb, SHEET_NAMES['score_psi'], score_psi_df)

    # 16_Bootstrap_CI
    if bootstrap_df is not None and len(bootstrap_df) > 0:
        _write_df_sheet(wb, SHEET_NAMES['bootstrap_ci'], bootstrap_df)

    # 17_SHAP
    if shap_summary_df is not None and len(shap_summary_df) > 0:
        _write_shap_sheet(wb, SHEET_NAMES['shap'], shap_summary_df, shap_plot_path)

    # 18_Calibration
    if calibration_dict is not None:
        _write_best_params_sheet_named(wb, SHEET_NAMES['calibration'], calibration_dict)

    # 19_Subseg_Perf
    if subsegment_perf:
        combined_rows = []
        for seg_col, seg_df in subsegment_perf.items():
            seg_copy = seg_df.copy()
            if 'Subsegment_Column' not in seg_copy.columns:
                seg_copy.insert(0, 'Subsegment_Column', seg_col)
            combined_rows.append(seg_copy)
        if combined_rows:
            combined = pd.concat(combined_rows, ignore_index=True)
            _write_df_sheet(wb, SHEET_NAMES['subseg_perf'], combined)

    # 20_Subseg_Confusion
    if subsegment_confusion:
        combined_rows = []
        for seg_col, seg_df in subsegment_confusion.items():
            seg_copy = seg_df.copy()
            if 'Subsegment_Column' not in seg_copy.columns:
                seg_copy.insert(0, 'Subsegment_Column', seg_col)
            combined_rows.append(seg_copy)
        if combined_rows:
            combined = pd.concat(combined_rows, ignore_index=True)
            _write_df_sheet(wb, SHEET_NAMES['subseg_confusion'], combined)

    # 21_Validation
    if validation_report_df is not None and len(validation_report_df) > 0:
        _write_validation_sheet(wb, SHEET_NAMES['validation'], validation_report_df)

    # Remove default empty sheet if exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Reorder sheets by their numerical prefix (e.g., 00, 01, ..., 10a, 10b, 11, 11a, 11b)
    def _sheet_sort_key(name):
        prefix = name.split('_')[0]
        # Extract numeric part and optional letter suffix
        num = ''.join(c for c in prefix if c.isdigit())
        suffix = ''.join(c for c in prefix if c.isalpha())
        return (int(num) if num else 0, suffix or '', name)

    sheet_order = sorted(wb.sheetnames, key=_sheet_sort_key)
    wb._sheets = [wb[name] for name in sheet_order]

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"COMPLETE | Excel saved: {output_path}")
    return output_path


def _write_summary_sheet(wb: Workbook, summary: Dict[str, Any]) -> None:
    """Write the 00_Summary sheet."""
    ws = wb.create_sheet("00_Summary")

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50

    # Title
    ws['A1'] = "Credit Scoring Model Development Report"
    ws['A1'].font = Font(bold=True, size=14, color="2F5496")
    ws.merge_cells('A1:B1')

    row = 3
    for key, value in summary.items():
        cell_a = ws.cell(row=row, column=1, value=key)
        cell_b = ws.cell(row=row, column=2, value=str(value))
        cell_a.font = Font(bold=True)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
        row += 1


def _write_selection_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    chart_path: Optional[str] = None,
) -> None:
    """Write the selection sheet with optional embedded chart."""
    if df is None or len(df) == 0:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = "No data"
        return

    ws = wb.create_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            if pd.isna(value):
                cell.value = None
            elif isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = float(value)
            elif isinstance(value, (bool, np.bool_)):
                cell.value = bool(value)
            else:
                cell.value = value

        # Highlight optimal row
        if 'Is_Optimal' in df.columns:
            opt_col = list(df.columns).index('Is_Optimal') + 1
            opt_val = ws.cell(row=row_idx, column=opt_col).value
            if opt_val is True or opt_val == 'True':
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = OPTIMAL_FILL

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for r in range(2, min(len(df) + 2, 102)):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 40)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Embed chart image if available
    if chart_path and Path(chart_path).exists():
        try:
            img = OpenpyxlImage(chart_path)
            img.width = 800
            img.height = 480
            # Place below the data table
            img_row = len(df) + 4
            ws.add_image(img, f'A{img_row}')
            logger.info(f"EXCEL | Embedded selection chart in {sheet_name}")
        except Exception as e:
            logger.warning(f"EXCEL | Could not embed chart: {e}")


def _write_best_params_sheet(
    wb: Workbook, best_params: Dict[str, Any]
) -> None:
    """Write the tuning best params sheet with best hyperparameters."""
    ws = wb.create_sheet(SHEET_NAMES['tuning_best'])

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25

    # Header
    for col_idx, col_name in enumerate(['Parameter', 'Value'], 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write params
    row = 2
    for key, value in sorted(best_params.items()):
        cell_a = ws.cell(row=row, column=1, value=str(key))
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=row, column=2)
        cell_b.border = THIN_BORDER
        if isinstance(value, (np.integer,)):
            cell_b.value = int(value)
        elif isinstance(value, (np.floating, float)):
            cell_b.value = round(float(value), 6)
        else:
            cell_b.value = str(value)
        row += 1

    ws.freeze_panes = 'A2'


def _write_df_sheet(
    wb: Workbook, sheet_name: str, df: pd.DataFrame
) -> None:
    """Write a DataFrame to a styled sheet."""
    if df is None or len(df) == 0:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = "No data"
        return

    ws = wb.create_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER

            # Handle numpy/pandas types
            if pd.isna(value):
                cell.value = None
            elif isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = float(value)
            else:
                cell.value = value

        # Color code rows by Status column if present
        if 'Status' in df.columns:
            status_col = list(df.columns).index('Status') + 1
            status_val = ws.cell(row=row_idx, column=status_col).value
            fill = None
            if status_val == 'Eliminated':
                fill = ELIM_FILL
            elif status_val == 'Kept' or status_val == 'Added':
                fill = KEPT_FILL

            if fill:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-fit column widths (approximate)
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 102)):  # sample first 100 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 40)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _write_lift_sheets(
    wb: Workbook, lift_tables: Dict[str, pd.DataFrame]
) -> None:
    """Write combined lift table sheet."""
    if not lift_tables:
        return

    combined_rows = []
    for period, lt in lift_tables.items():
        lt_copy = lt.copy()
        lt_copy.insert(0, 'Period', period)
        combined_rows.append(lt_copy)

    if combined_rows:
        combined = pd.concat(combined_rows, ignore_index=True)
        _write_df_sheet(wb, SHEET_NAMES['lift_tables'], combined)


# Styles for validation
PASS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")


def _write_shap_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    plot_path: Optional[str] = None,
) -> None:
    """Write SHAP summary sheet with optional embedded plot."""
    if df is None or len(df) == 0:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = "No data"
        return

    ws = wb.create_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if pd.isna(value):
                cell.value = None
            elif isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = float(value)
            else:
                cell.value = value

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for r in range(2, min(len(df) + 2, 102)):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 40)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Embed SHAP plot if available
    if plot_path and Path(plot_path).exists():
        try:
            img = OpenpyxlImage(plot_path)
            img.width = 800
            img.height = 480
            img_row = len(df) + 4
            ws.add_image(img, f'A{img_row}')
            logger.info(f"EXCEL | Embedded SHAP plot in {sheet_name}")
        except Exception as e:
            logger.warning(f"EXCEL | Could not embed SHAP plot: {e}")


def _write_best_params_sheet_named(
    wb: Workbook, sheet_name: str, params: Dict[str, Any]
) -> None:
    """Write a key-value sheet (like calibration results)."""
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25

    # Header
    for col_idx, col_name in enumerate(['Parameter', 'Value'], 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    row = 2
    for key, value in sorted(params.items()):
        cell_a = ws.cell(row=row, column=1, value=str(key))
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=row, column=2)
        cell_b.border = THIN_BORDER
        if isinstance(value, (np.integer,)):
            cell_b.value = int(value)
        elif isinstance(value, (np.floating, float)):
            cell_b.value = round(float(value), 6)
        else:
            cell_b.value = str(value)
        row += 1

    ws.freeze_panes = 'A2'


def _write_validation_sheet(
    wb: Workbook, sheet_name: str, df: pd.DataFrame
) -> None:
    """Write validation report with color-coded status (PASS=green, FAIL=red, WARNING=yellow)."""
    if df is None or len(df) == 0:
        ws = wb.create_sheet(sheet_name)
        ws['A1'] = "No data"
        return

    ws = wb.create_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

    # Write data rows with color coding
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        status_val = None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if pd.isna(value):
                cell.value = None
            elif isinstance(value, (np.integer,)):
                cell.value = int(value)
            elif isinstance(value, (np.floating,)):
                cell.value = float(value)
            else:
                cell.value = value

            # Track status for row coloring
            if df.columns[col_idx - 1] == 'Status':
                status_val = value

        # Color entire row based on status
        if status_val:
            fill = None
            if status_val == 'PASS':
                fill = PASS_FILL
            elif status_val == 'FAIL':
                fill = FAIL_FILL
            elif status_val == 'WARNING':
                fill = WARNING_FILL
            if fill:
                for col_idx in range(1, len(df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for r in range(2, min(len(df) + 2, 102)):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 50)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
