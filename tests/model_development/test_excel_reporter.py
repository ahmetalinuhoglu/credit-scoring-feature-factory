"""
Tests for src.model_development.excel_reporter

Covers: generate_report, SHEET_NAMES, required/optional sheets.
"""

import pytest
import numpy as np
import pandas as pd
from pathlib import Path

from src.model_development.excel_reporter import generate_report, SHEET_NAMES
from src.model_development.eliminators import EliminationResult


# ===================================================================
# Helper to build minimal report inputs
# ===================================================================

def _make_report_inputs():
    """Build minimal valid inputs for generate_report."""
    summary = {
        "Run Date": "2024-01-01",
        "Run ID": "test_run",
        "Total Features": 10,
    }

    details_df = pd.DataFrame({
        "Feature": ["f1", "f2"],
        "Status": ["Kept", "Eliminated"],
    })
    elim_results = [
        EliminationResult("01_Constant", ["f1"], ["f2"], details_df),
    ]

    selection_df = pd.DataFrame({
        "Step": [1],
        "N_Features": [1],
        "Added_Feature": ["f1"],
        "Mean_CV_AUC": [0.80],
        "Std_CV_AUC": [0.01],
        "Is_Optimal": [True],
    })

    performance_df = pd.DataFrame({
        "Period": ["Train", "Test"],
        "N_Samples": [80, 20],
        "N_Bads": [16, 4],
        "Bad_Rate": [0.20, 0.20],
        "AUC": [0.85, 0.82],
        "Gini": [0.70, 0.64],
        "KS": [0.55, 0.52],
    })

    lift_df = pd.DataFrame({
        "decile": [1, 2],
        "Count": [10, 10],
        "Bads": [5, 3],
        "Bad_Rate": [0.50, 0.30],
    })
    lift_tables = {"Train": lift_df}

    importance_df = pd.DataFrame({
        "Feature": ["f1"],
        "Importance": [1.0],
        "Rank": [1],
        "Cumulative_Importance": [1.0],
    })

    return {
        "summary": summary,
        "elimination_results": elim_results,
        "corr_pairs_df": None,
        "selection_df": selection_df,
        "performance_df": performance_df,
        "lift_tables": lift_tables,
        "importance_df": importance_df,
    }


# ===================================================================
# Tests
# ===================================================================

class TestGenerateReport:
    def test_produces_xlsx(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        result = generate_report(output_path=path, **inputs)
        assert Path(result).exists()
        assert result.endswith(".xlsx")

    def test_file_is_valid_xlsx(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert len(wb.sheetnames) > 0

    def test_summary_sheet_present(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert "00_Summary" in wb.sheetnames

    def test_performance_sheet_present(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert SHEET_NAMES["performance"] in wb.sheetnames

    def test_importance_sheet_present(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert SHEET_NAMES["importance"] in wb.sheetnames


class TestSheetNames:
    def test_has_expected_keys(self):
        expected_keys = [
            "summary", "constant", "missing", "iv", "psi",
            "correlation", "corr_pairs", "selection",
            "performance", "lift_tables", "importance",
        ]
        for key in expected_keys:
            assert key in SHEET_NAMES

    def test_values_are_strings(self):
        for key, val in SHEET_NAMES.items():
            assert isinstance(val, str)


class TestOptionalSheets:
    def test_tuning_omitted_when_none(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, tuning_df=None, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert SHEET_NAMES["tuning"] not in wb.sheetnames

    def test_tuning_present_when_provided(self, tmp_path):
        inputs = _make_report_inputs()
        tuning_df = pd.DataFrame({
            "Trial": [0, 1],
            "max_depth": [3, 4],
            "CV_AUC_Mean": [0.80, 0.82],
        })
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, tuning_df=tuning_df, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert SHEET_NAMES["tuning"] in wb.sheetnames

    def test_bootstrap_omitted_when_none(self, tmp_path):
        inputs = _make_report_inputs()
        path = str(tmp_path / "report.xlsx")
        generate_report(output_path=path, bootstrap_df=None, **inputs)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert SHEET_NAMES["bootstrap_ci"] not in wb.sheetnames
