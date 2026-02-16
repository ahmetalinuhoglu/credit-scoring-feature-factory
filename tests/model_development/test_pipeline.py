"""
Tests for src.model_development.pipeline

Covers: end-to-end pipeline run with small data, _build_summary.
"""

import pytest
import numpy as np
import pandas as pd
from unittest.mock import patch, MagicMock
from pathlib import Path

from src.model_development.pipeline import ModelDevelopmentPipeline


# ===================================================================
# Helper to build small parquet for pipeline
# ===================================================================

def _make_pipeline_parquet(tmp_path, n=300, seed=42):
    """Write a small parquet for end-to-end pipeline testing.

    Uses n=300 to ensure enough rows survive train/test split for
    IV calculation (needs >= 50 rows), and enough quarters for PSI.
    """
    rng = np.random.RandomState(seed)
    target = np.zeros(n, dtype=int)
    target[: int(n * 0.20)] = 1
    rng.shuffle(target)

    dates = pd.date_range("2024-01-01", "2024-09-30", periods=n)

    data = {
        "application_id": range(n),
        "customer_id": range(1000, 1000 + n),
        "applicant_type": rng.choice(["new", "existing"], size=n),
        "application_date": dates,
        "target": target,
    }
    # 5 features with moderate signal (not too strong to avoid suspicious IV)
    for i in range(5):
        strength = 0.3 + i * 0.1
        data[f"feat_{i}"] = target * strength + rng.randn(n)

    df = pd.DataFrame(data)
    path = str(tmp_path / "pipeline_data.parquet")
    df.to_parquet(path, index=False)
    return path


# ===================================================================
# End-to-end test
# ===================================================================

class TestPipelineEndToEnd:
    def test_runs_without_error(self, tmp_path):
        """Quick end-to-end: no tuning, no VIF, small data."""
        path = _make_pipeline_parquet(tmp_path)
        output_dir = str(tmp_path / "output")

        # Mock PipelineConfig to avoid needing the real config schema
        mock_config = MagicMock()
        mock_config.reproducibility.n_jobs = 1
        mock_config.splitting.stratify = True
        mock_config.steps.temporal_filter.enabled = False
        mock_config.evaluation = None
        mock_config.validation.enabled = False
        mock_config.output.save_model = False

        pipeline = ModelDevelopmentPipeline(
            input_path=path,
            train_end_date="2024-06-30",
            output_dir=output_dir,
            tuning_enabled=False,
            vif_enabled=False,
            iv_max=10.0,  # Wide IV bounds for small synthetic data
            selection_max_features=3,
            selection_cv=2,
            selection_patience=1,
            config=mock_config,
        )
        results = pipeline.run()
        assert results["status"] == "success"
        assert Path(results["excel_path"]).exists()

    def test_selected_features_in_results(self, tmp_path):
        path = _make_pipeline_parquet(tmp_path)
        output_dir = str(tmp_path / "output")

        mock_config = MagicMock()
        mock_config.reproducibility.n_jobs = 1
        mock_config.splitting.stratify = True
        mock_config.steps.temporal_filter.enabled = False
        mock_config.evaluation = None
        mock_config.validation.enabled = False
        mock_config.output.save_model = False

        pipeline = ModelDevelopmentPipeline(
            input_path=path,
            train_end_date="2024-06-30",
            output_dir=output_dir,
            tuning_enabled=False,
            vif_enabled=False,
            iv_max=10.0,
            selection_max_features=3,
            selection_cv=2,
            selection_patience=1,
            config=mock_config,
        )
        results = pipeline.run()
        assert "selected_features" in results
        assert len(results["selected_features"]) >= 1

    def test_excel_created(self, tmp_path):
        path = _make_pipeline_parquet(tmp_path)
        output_dir = str(tmp_path / "output")

        mock_config = MagicMock()
        mock_config.reproducibility.n_jobs = 1
        mock_config.splitting.stratify = True
        mock_config.steps.temporal_filter.enabled = False
        mock_config.evaluation = None
        mock_config.validation.enabled = False
        mock_config.output.save_model = False

        pipeline = ModelDevelopmentPipeline(
            input_path=path,
            train_end_date="2024-06-30",
            output_dir=output_dir,
            tuning_enabled=False,
            vif_enabled=False,
            iv_max=10.0,
            selection_max_features=2,
            selection_cv=2,
            selection_patience=1,
            config=mock_config,
        )
        results = pipeline.run()
        excel_path = results["excel_path"]
        assert Path(excel_path).exists()
        assert excel_path.endswith(".xlsx")


# ===================================================================
# _build_summary test
# ===================================================================

class TestBuildSummary:
    def test_returns_dict_with_expected_keys(self, tmp_path):
        path = _make_pipeline_parquet(tmp_path)
        output_dir = str(tmp_path / "output")

        mock_config = MagicMock()
        mock_config.reproducibility.n_jobs = 1
        mock_config.splitting.stratify = True
        mock_config.steps.temporal_filter.enabled = False
        mock_config.evaluation = None
        mock_config.validation.enabled = False
        mock_config.output.save_model = False

        pipeline = ModelDevelopmentPipeline(
            input_path=path,
            train_end_date="2024-06-30",
            output_dir=output_dir,
            tuning_enabled=False,
            vif_enabled=False,
            iv_max=10.0,
            selection_max_features=2,
            selection_cv=2,
            selection_patience=1,
            config=mock_config,
        )
        results = pipeline.run()

        # We cannot call _build_summary directly in a clean way
        # without running the pipeline, but we can verify the
        # summary ended up in the Excel
        from openpyxl import load_workbook
        wb = load_workbook(results["excel_path"])
        assert "00_Summary" in wb.sheetnames
        ws = wb["00_Summary"]
        # Check that some expected keys are in column A
        col_a_values = [ws.cell(row=r, column=1).value for r in range(1, 40)]
        assert "Run ID" in col_a_values
        assert "Total Features" in col_a_values

    def test_summary_has_feature_counts(self, tmp_path):
        path = _make_pipeline_parquet(tmp_path)
        output_dir = str(tmp_path / "output")

        mock_config = MagicMock()
        mock_config.reproducibility.n_jobs = 1
        mock_config.splitting.stratify = True
        mock_config.steps.temporal_filter.enabled = False
        mock_config.evaluation = None
        mock_config.validation.enabled = False
        mock_config.output.save_model = False

        pipeline = ModelDevelopmentPipeline(
            input_path=path,
            train_end_date="2024-06-30",
            output_dir=output_dir,
            tuning_enabled=False,
            vif_enabled=False,
            iv_max=10.0,
            selection_max_features=2,
            selection_cv=2,
            selection_patience=1,
            config=mock_config,
        )
        results = pipeline.run()

        # Check that pipeline counts are present
        assert "total_features" in results
        assert "after_constant" in results
        assert "after_selection" in results
